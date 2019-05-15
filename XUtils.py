import openpyxl


def getRowCount(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return sheet.max_row

def getColumnCount(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return sheet.max_column

def readData(file, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    return sheet.cell(row= rownum, column= columnnum).value

def writeData (file, rows, columns, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(row= rows, column= columns).value = data
    workbook.save(file)


