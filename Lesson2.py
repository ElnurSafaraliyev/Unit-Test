import openpyxl

path = "C:/Users/Cissy/Desktop/Canadian All Care/QA/Testing/DDT/DATA2.xlsx"


workbook = openpyxl.load_workbook(path)
sheet = workbook.active



for r in range(1, 6):
    for c in range (1, 4):
        sheet.cell(row= r, column= c).value = "Welcome"
workbook.save(path)


sheet.cell(row=10, column= 10).value = 10+10
workbook.save(path)


